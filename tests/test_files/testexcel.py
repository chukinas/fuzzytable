from openpyxl import load_workbook
from fuzzytable.main.sheetreader import ExcelReader



path = 'test.xlsx'
sheetname = 'table_top_left'
sheetreader = ExcelReader(path, sheetname)


# def get_reader():
#     wb = load_workbook(path, read_only=True)  # Lazy loader
#     ws = wb[sheetname]
#     yield ws.iter_rows(values_only=True)


row = sheetreader[1]
print('row')
print(type(row))
print(row)

print(sheetreader.get_col(3))



# rows = get_reader()
# print('rows')
# print(type(rows))
# print(rows)
# for row in rows:
#     print('row')
#     print(type(row))
#     print(row)
#     for cell in row:
#         print('cell')
#         print(type(cell))
#         print(cell)

# row = [cell.value for cell in row]
