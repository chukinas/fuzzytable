"""
Interface between the csv/excel files and the rest of FuzzyTable
"""

# --- Standard Library Imports ------------------------------------------------
import csv
from contextlib import contextmanager
from typing import List
from ast import literal_eval

# --- Intra-Package Imports ---------------------------------------------------
from fuzzytable import exceptions

# --- Third Party Imports -----------------------------------------------------
from openpyxl.worksheet.worksheet import Worksheet as openpyxlWorksheet
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


INFINITY = float("inf")
NEG_INFINITY = float("-inf")


class SheetReader:

    def __init__(self, path, sheetname=None) -> None:
        self.path = path
        self._row_count = None
        self.sheetname = sheetname

    def iter_row(self, start_row=None, end_row=None):
        # Generator Function for looping over rows
        if start_row is None:
            start_row = NEG_INFINITY
        if end_row is None:
            end_row = INFINITY
        with self.get_filereader() as filereader:
            for row_num, row in enumerate(filereader, 1):
                if row_num > end_row:
                    return
                if row_num >= start_row:
                    yield row
        self._row_count = row_num

    def iter_row_str(self, end_row=None):
        # Generator function for looping over row repr's
        for row in self.iter_row(end_row=end_row):
            yield repr(row)

    def get_row_str(self, row_num):
        # Use this only when you only care about this row and no other.
        # When looping to find best fit, use self.iter_row_str instead.
        row = self[row_num]
        return repr(row)

    def get_col(self, col_num, start_row=1, cellpatterns=None):
        # Return column values as list
        col_index = col_num - 1
        values = [row[col_index] for row in self.iter_row(start_row=start_row)]
        cellpatterns = force_list(cellpatterns)
        cellpatterns.insert(0, _eval)
        for cellpattern in cellpatterns:
            for index, value in enumerate(values):
                new_val = cellpattern(value)
                values[index] = new_val
            # NOTE: the above for loop replace the below comprehension for debugging purposes.
            # I plan to change it back eventually
            # values = [cellpattern(value) for value in values]
        return values

    @property
    def row_count(self):
        if self._row_count is None:
            for _ in self.iter_row():
                pass  # iter_row populates self._row_count at the end
            return self.row_count
        else:
            return self._row_count

    def __getitem__(self, desired_row_num):
        for cur_row_num, row in enumerate(self.iter_row(), 1):
            if cur_row_num == desired_row_num:
                return row

    @contextmanager
    def get_filereader(self):
        file = open(self.path)
        yield csv.reader(file)
        file.close()

    # def __repr__(self):
    #     return get_repr(self)  # pragma: no cover


class CsvReader(SheetReader):
    # def get_col(self, start_row, col_num, cellpatterns=None):
    #     cellpatterns = force_list(cellpatterns)
    #     # cellpatterns.insert(0, _eval)
    #     return super().get_col(start_row=start_row, col_num=col_num, cellpatterns=cellpatterns)
    pass

def force_list(value) -> List:
    if value is None:
        return []
    try:
        return list(value)
    except TypeError:
        return [value]


class ExcelReader(SheetReader):

    @contextmanager
    def get_filereader(self):
        try:
            wb = load_workbook(self.path, read_only=True)  # Lazy loader
            ws: openpyxlWorksheet = wb[self.sheetname]
        except InvalidFileException:
            raise exceptions.InvalidFileError(self.path)
        except KeyError:
            # worksheet not found
            raise exceptions.SheetnameError(self.path, self.sheetname)
        yield ws.iter_rows(values_only=True)
        # row = [cell.value for cell in row]
        # yield row

    # def get_col(self, start_row, col_num):
    #     orig_col = super().get_col(start_row=start_row, col_num=col_num)
    #     return [str(val) for val in orig_col]


def _eval(value):
    if value == '':
        return None
    try:
        return literal_eval(value)
    except (SyntaxError, ValueError):
        return _bool(value)


def _bool(value):
    if isinstance(value, bool):
        return value
    elif value in 'TRUE True true'.split():
        return True
    elif value in 'FALSE false False'.split():
        return False
    else:
        return value

if __name__ == '__main__':
    pass
