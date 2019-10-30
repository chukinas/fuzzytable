""""""
# TODO Fill in docstring

# --- Standard Library Imports ------------------------------------------------
# None

# --- Third Party Imports -----------------------------------------------------
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# --- Intra-Package Imports ---------------------------------------------------
from excelerator import exceptions


def get_workbook(path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except (FileNotFoundError, InvalidFileException):
        raise exceptions.EXCELeratorError(f"\nError: '{path} ' is not a valid excel path.")


def get_worksheet_from_workbook(workbook, worksheet_name):
    try:
        return workbook[worksheet_name]
    except KeyError:
        raise exceptions.EXCELeratorError(f"\nError: Workbook does not contain a '{worksheet_name}' worksheet")


def get_worksheet_from_path(path, sheetname):
    workbook = get_workbook(path)
    return get_worksheet_from_workbook(workbook, sheetname)


def get_column(worksheet, col_num, row_start, row_end):
    return [
        worksheet.cell(row, col_num).value
        for row in range(row_start, row_end + 1)
    ]


def get_worksheet_row(
        worksheet,
        row_int,
        # drop_blanks=False,
        # dups_raise_error=False,
        # rowval_raise_error=False,
) -> list:
    """

    :param worksheet: openpyxl Worksheet object
    :param row_int:
    :param drop_blanks:
    :param dups_raise_error:
    :param rowval_raise_error: Often, you want to create a custom field called 'row' to save the row number.
    :return:
    """

    # --- get full row --------------------------------------------------------
    row = [worksheet.cell(row_int, col).value for col in range(1, worksheet.max_column + 1)]

    # --- remove blanks (if desired) ------------------------------------------
    # if drop_blanks:
    #     row = [value for value in row if value is not None]

    # --- check for duplicates (if desired) -----------------------------------
    # if dups_raise_error:
    #     value_counts = collections.Counter(row)
    #     repeated_values = [header for header in row if value_counts[header] > 1]
    #     if repeated_values:
    #         msg = f'\nError: the following headers in {worksheet.title} have duplicates:\n{repeated_values}'
    #         raise exceptions.ExcelUtilError(msg)

    # --- check for value == 'row' (if desired) -------------------------------
    # if rowval_raise_error and 'row' in row:
    #     msg = f"\nError: 'row' is not a valid header in {worksheet.title}"
    #     raise exceptions.ExcelUtilError(msg)
    return row


# def get_best_match_row_number(worksheet, values, max_row=20):
#     """Find the row that best matches a set of values.
#     :param worksheet: openpyxl Worksheet object
#     :param values: collection of row values you expect to find.
#     :param max_row: function searches rows from 1 to max_row
#     :return: row number (1-indexed) that best matches the provided values
#     """
#     rows = [get_worksheet_row(worksheet, row) for row in range(1, max_row + 1)]
#     best_index = string_analysis.find_most_similar_string_sequence(values, rows)
#     return best_index + 1
